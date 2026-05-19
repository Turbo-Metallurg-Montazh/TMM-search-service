import os
import glob
from typing import List, Tuple

import numpy as np
import pandas as pd
import torch
import torch.nn.functional as F
from transformers import AutoTokenizer, AutoModel

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import math

# ---------------- CONFIG ----------------

SAMPLES_FILE = "samples.xlsx"        # input file with Column A and (optionally) Column B
DAPT_DIR = "dapt"                    # directory with .xlsx/.xls files used for DAPT
MODEL_DIR = "biencoder_model"        # your trained bi-encoder model

MAX_LEN = 128
ENC_BATCH_SIZE = 64                  # batch size for encoding candidates & queries

# threshold for "closely matches" between column A and E (reference B)
AE_SIM_THRESHOLD = 0.8

DEVICE = (
    "mps" if torch.backends.mps.is_available()
    else "cuda" if torch.cuda.is_available()
    else "cpu"
)

# ---------------- CANDIDATE LOADING ----------------

def extract_texts_from_excel_file(path: str) -> List[str]:
    """
    Extract all non-empty cell texts from all sheets in a single Excel file.
    Returns a list of strings.
    """
    texts = []
    try:
        xls = pd.ExcelFile(path)
        for sheet in xls.sheet_names:
            df = pd.read_excel(path, sheet_name=sheet, dtype=str)
            # flatten all cell values
            for col in df.columns:
                col_values = df[col].dropna().astype(str)
                for v in col_values:
                    v_strip = v.strip()
                    if v_strip:
                        texts.append(v_strip)
    except Exception as e:
        print(f"Failed to read Excel file {path}: {e}")
    return texts


def load_all_candidates_from_dapt(dapt_dir: str) -> List[str]:
    """
    Load all candidate texts from .xlsx/.xls files inside dapt_dir (recursively).
    Returns a deduplicated list of non-empty strings.
    """
    candidates = []
    pattern = os.path.join(dapt_dir, "**", "*")
    for path in glob.glob(pattern, recursive=True):
        if os.path.isdir(path):
            continue
        lower = path.lower()
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            print(f"Reading candidates from Excel file: {path}")
            candidates.extend(extract_texts_from_excel_file(path))

    # Deduplicate while preserving order
    seen = set()
    unique_candidates = []
    for t in candidates:
        if t not in seen:
            seen.add(t)
            unique_candidates.append(t)

    print(f"Total candidate strings (deduplicated): {len(unique_candidates)}")
    return unique_candidates


# ---------------- MODEL ENCODING ----------------

def encode_texts(texts: List[str],
                 tokenizer: AutoTokenizer,
                 encoder: AutoModel,
                 device: str,
                 max_len: int = MAX_LEN,
                 batch_size: int = ENC_BATCH_SIZE) -> torch.Tensor:
    """
    Encode a list of texts into L2-normalized embeddings using the biencoder encoder.
    Returns a tensor of shape (N, H) on the specified device.
    """
    all_embeddings = []

    for i in range(0, len(texts), batch_size):
        batch_texts = texts[i:i + batch_size]
        enc = tokenizer(
            batch_texts,
            padding=True,
            truncation=True,
            max_length=max_len,
            return_tensors="pt"
        )
        input_ids = enc["input_ids"].to(device)
        attention_mask = enc["attention_mask"].to(device)

        with torch.no_grad():
            outputs = encoder(input_ids=input_ids, attention_mask=attention_mask)
            last_hidden = outputs.last_hidden_state  # (B, T, H)
            mask = attention_mask.unsqueeze(-1)      # (B, T, 1)
            masked = last_hidden * mask
            summed = masked.sum(dim=1)
            counts = mask.sum(dim=1).clamp(min=1e-9)
            mean = summed / counts                   # (B, H)
            normed = F.normalize(mean, p=2, dim=1)   # L2-normalized

        all_embeddings.append(normed.cpu())

    if not all_embeddings:
        return torch.empty(0, encoder.config.hidden_size, device=device)

    embs = torch.cat(all_embeddings, dim=0).to(device)
    return embs


# ---------------- MATCHING LOGIC ----------------

def find_best_two(
    query_emb: torch.Tensor,
    candidate_embs: torch.Tensor
) -> Tuple[int, int, float, float]:
    """
    Given query_emb of shape (H,) and candidate_embs of shape (N, H),
    compute cosine similarities and return:
    (best_idx, second_idx, best_score, second_score).
    """
    with torch.no_grad():
        q = query_emb.unsqueeze(0)  # (1, H)
        sims = torch.matmul(q, candidate_embs.t()).cpu().numpy()[0]  # (N,)

    if len(sims) == 0:
        return -1, -1, float("nan"), float("nan")

    sorted_idx = np.argsort(-sims)  # descending
    best_idx = int(sorted_idx[0])
    best_score = float(sims[best_idx])

    if len(sorted_idx) > 1:
        second_idx = int(sorted_idx[1])
        second_score = float(sims[second_idx])
    else:
        second_idx = -1
        second_score = float("nan")

    return best_idx, second_idx, best_score, second_score


# ---------------- MAIN PIPELINE ----------------

def main():
    print(f"DEVICE: {DEVICE}")

    # 1. Load model + tokenizer
    print(f"Loading biencoder model from: {MODEL_DIR}")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_DIR)
    encoder = AutoModel.from_pretrained(MODEL_DIR).to(DEVICE)
    encoder.eval()

    # 2. Load candidates from dapt Excel files
    candidates = load_all_candidates_from_dapt(DAPT_DIR)
    if not candidates:
        raise RuntimeError("No candidates found in dapt directory (.xlsx/.xls).")

    print("Encoding candidate texts...")
    candidate_embs = encode_texts(candidates, tokenizer, encoder, DEVICE)
    print(f"Candidate embeddings shape: {candidate_embs.shape}")

    # 3. Load samples.xlsx
    print(f"Loading samples from: {SAMPLES_FILE}")
    df_samples = pd.read_excel(SAMPLES_FILE)

    # assume Column A is first column; Column B (if exists) is reference
    col_a = df_samples.iloc[:, 0].astype(str).tolist()
    if df_samples.shape[1] >= 2:
        col_b_ref = df_samples.iloc[:, 1].astype(str).tolist()
    else:
        col_b_ref = [""] * len(col_a)

    print(f"Number of sample rows: {len(col_a)}")

    # 4. Encode all sample A texts AND reference B (Column E) texts
    print("Encoding sample A texts...")
    sample_embs = encode_texts(col_a, tokenizer, encoder, DEVICE)
    print(f"Sample embeddings shape: {sample_embs.shape}")

    print("Encoding reference B (Column E) texts...")
    ref_embs = encode_texts(col_b_ref, tokenizer, encoder, DEVICE)
    print(f"Reference embeddings shape: {ref_embs.shape}")

    # 5. For each sample, find best & second-best candidate, and A–E similarity
    out_rows = []
    ae_sims = []  # to use for coloring later

    for i, (text_a, ref_b) in enumerate(zip(col_a, col_b_ref)):
        query_emb = sample_embs[i]  # (H,)
        best_idx, second_idx, best_score, second_score = find_best_two(query_emb, candidate_embs)

        if best_idx == -1:
            best_match = ""
            second_match = ""
            score_diff = float("nan")
            best_match_sim = float("nan")
        else:
            best_match = candidates[best_idx]
            best_match_sim = best_score  # similarity between A and best_match

            if second_idx == -1:
                second_match = ""
                score_diff = float("nan")
            else:
                second_match = candidates[second_idx]
                score_diff = best_score - second_score

        # A–E similarity (A_sample vs reference B)
        if ref_b.strip() == "":
            ae_sim = float("nan")
        else:
            with torch.no_grad():
                ae_sim = float(torch.dot(sample_embs[i], ref_embs[i]).item())

        ae_sims.append(ae_sim)

        out_rows.append({
            "A_sample": text_a,                 # Column A
            "B_best_match": best_match,         # Column B
            "C_second_match": second_match,     # Column C
            "D_best_match_similarity": best_match_sim,  # NEW Column D
            "E_score_diff": score_diff,         # Column E (was D)
            "F_sample_B_ref": ref_b,            # Column F
            "G_AE_similarity": ae_sim           # Column G
        })

        if (i + 1) % 50 == 0:
            print(f"Processed {i+1}/{len(col_a)} rows...")

    # 6. Save to new Excel file
    out_df = pd.DataFrame(out_rows, columns=[
        "A_sample",
        "B_best_match",
        "C_second_match",
        "D_best_match_similarity",
        "E_score_diff",
        "F_sample_B_ref",
        "G_AE_similarity",
    ])

    output_file = "samples_with_matches.xlsx"
    out_df.to_excel(output_file, index=False)
    print(f"Results saved (before coloring) to: {output_file}")

    # 7. Color rows green where A closely matches E (similarity above threshold)
    wb = load_workbook(output_file)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Row 1 is header; data starts at row 2
    for row_idx, sim in enumerate(ae_sims, start=2):
        if sim is None or math.isnan(sim):
            continue
        if sim >= AE_SIM_THRESHOLD:
            for cell in ws[row_idx]:
                cell.fill = green_fill

    wb.save(output_file)
    print(f"Applied green highlighting (A–E sim >= {AE_SIM_THRESHOLD}) and saved to: {output_file}")


if __name__ == "__main__":
    main()
