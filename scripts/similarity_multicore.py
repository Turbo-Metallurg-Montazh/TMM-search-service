import os
import glob
import math
import re
from typing import List, Tuple, Optional

import numpy as np
import pandas as pd
import torch
import torch.nn.functional as F
from transformers import AutoTokenizer, AutoModel

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ---------------- CONFIG ----------------

SAMPLES_FILE = "samplescleaned.xlsx"
DAPT_DIR = "dapt"
MODEL_DIR = "biencoder_model"

MAX_LEN = 128
ENC_BATCH_SIZE = 64

AE_SIM_THRESHOLD = 0.8

GOOD_SIM_THRESHOLD = 0.85          # your requirement
MISSING_PRICE = 1e18               # treat missing price as extremely expensive

DEVICE = (
    "mps" if torch.backends.mps.is_available()
    else "cuda" if torch.cuda.is_available()
    else "cpu"
)


# ---------------- PARAMETER PARSING (same as before) ----------------

SOCKET_RE = re.compile(r'\b(GU10|G[0-9]{2}|E[0-9]{2})\b', re.IGNORECASE)
TUBE_RE   = re.compile(r'\bT\s*([0-9])\b', re.IGNORECASE)
IP_RE     = re.compile(r'\bIP\s*([0-9]{2})\b', re.IGNORECASE)
VOLT_RE   = re.compile(r'(\d{2,4})\s*[VВ]\b', re.IGNORECASE)
AMP_RE    = re.compile(r'(\d{1,3})\s*[AА]\b', re.IGNORECASE)
POWER_RE  = re.compile(r'(\d{1,3})\s*(Вт|W)\b', re.IGNORECASE)
LENGTH_RE = re.compile(r'(\d{3,4})\s*(мм|mm)\b', re.IGNORECASE)
KELVIN_RE = re.compile(r'(\d{3,5})\s*[KК]\b', re.IGNORECASE)

def extract_params(text: str) -> dict:
    t = str(text).replace(",", " ")
    params = {}

    m = SOCKET_RE.search(t)
    if m: params["socket"] = m.group(1).upper()

    m = TUBE_RE.search(t)
    if m:
        try: params["tube"] = int(m.group(1))
        except: pass

    m = IP_RE.search(t)
    if m:
        try: params["ip"] = int(m.group(1))
        except: pass

    m = VOLT_RE.search(t)
    if m:
        try: params["volt"] = int(m.group(1))
        except: pass

    m = AMP_RE.search(t)
    if m:
        try: params["amp"] = int(m.group(1))
        except: pass

    m = POWER_RE.search(t)
    if m:
        try: params["watt"] = int(m.group(1))
        except: pass

    m = LENGTH_RE.search(t)
    if m:
        try: params["len"] = int(m.group(1))
        except: pass

    m = KELVIN_RE.search(t)
    if m:
        try: params["kelvin"] = int(m.group(1))
        except: pass

    return params

def relative_diff(a: float, b: float) -> float:
    m = max(abs(a), abs(b), 1e-9)
    return abs(a - b) / m

def parameter_mismatch_penalty(p_q: dict, p_c: dict) -> float:
    penalty = 0.0

    if "socket" in p_q and "socket" in p_c and p_q["socket"] != p_c["socket"]:
        penalty += 0.7
    if "tube" in p_q and "tube" in p_c and p_q["tube"] != p_c["tube"]:
        penalty += 0.5
    if "ip" in p_q and "ip" in p_c and p_q["ip"] != p_c["ip"]:
        penalty += 0.5

    if "volt" in p_q and "volt" in p_c and abs(p_q["volt"] - p_c["volt"]) >= 50:
        penalty += 0.4

    if "amp" in p_q and "amp" in p_c:
        if relative_diff(p_q["amp"], p_c["amp"]) >= 0.25:
            penalty += 0.3

    if "watt" in p_q and "watt" in p_c:
        rdiff = relative_diff(p_q["watt"], p_c["watt"])
        if rdiff >= 0.5: penalty += 0.3
        elif rdiff >= 0.2: penalty += 0.15

    if "len" in p_q and "len" in p_c:
        rdiff = relative_diff(p_q["len"], p_c["len"])
        if rdiff >= 0.3: penalty += 0.3
        elif rdiff >= 0.15: penalty += 0.15

    if "kelvin" in p_q and "kelvin" in p_c:
        diff_k = abs(p_q["kelvin"] - p_c["kelvin"])
        if diff_k >= 2000: penalty += 0.3
        elif diff_k >= 1000: penalty += 0.15

    return min(penalty, 1.5)


# ---------------- PRICE PARSING ----------------

def parse_price(x) -> Optional[float]:
    """
    Parse price from Excel cell.
    Accepts numbers or strings like "1 234,50" / "1234.50" / "1,234.50".
    Returns float or None.
    """
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return None
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        return float(x)

    s = str(x).strip()
    if not s:
        return None

    # keep digits, separators
    s = re.sub(r"[^\d.,]", "", s)
    if not s:
        return None

    # If both comma and dot exist, assume the last separator is decimal, remove the other as thousand-sep.
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        # If only comma exists, treat it as decimal if it looks like decimal, else thousand sep
        if "," in s and "." not in s:
            # e.g. "1234,56" -> decimal
            if re.search(r",\d{1,2}$", s):
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")

    # remove spaces just in case
    s = s.replace(" ", "")
    try:
        return float(s)
    except:
        return None


# ---------------- CANDIDATE LOADING (TEXT + PRICE) ----------------

def load_candidates_with_prices_from_excel(path: str) -> List[Tuple[str, Optional[float]]]:
    """
    Reads ALL sheets. Expects:
      - first column: candidate text
      - second column: price
    Returns list of (text, price).
    """
    out = []
    try:
        xls = pd.ExcelFile(path)
        for sheet in xls.sheet_names:
            df = pd.read_excel(path, sheet_name=sheet, dtype=object)
            if df.shape[1] < 1:
                continue
            text_col = df.iloc[:, 0]
            price_col = df.iloc[:, 1] if df.shape[1] >= 2 else None

            for idx in range(len(df)):
                t = "" if pd.isna(text_col.iloc[idx]) else str(text_col.iloc[idx]).strip()
                if not t:
                    continue
                p = parse_price(price_col.iloc[idx]) if price_col is not None else None
                out.append((t, p))
    except Exception as e:
        print(f"Failed to read Excel file {path}: {e}")
    return out

def load_all_candidates_from_dapt(dapt_dir: str):
    """
    Returns:
      candidates_texts: List[str]
      candidates_prices: List[Optional[float]]
    Dedup by text (keep the LOWEST price if duplicates).
    """
    rows = []
    pattern = os.path.join(dapt_dir, "**", "*")
    for path in glob.glob(pattern, recursive=True):
        if os.path.isdir(path):
            continue
        lower = path.lower()
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            print(f"Reading candidates from: {path}")
            rows.extend(load_candidates_with_prices_from_excel(path))

    # Deduplicate by text; keep min price
    best_price_by_text = {}
    for t, p in rows:
        if t not in best_price_by_text:
            best_price_by_text[t] = p
        else:
            old = best_price_by_text[t]
            # keep cheaper if both exist
            if old is None and p is not None:
                best_price_by_text[t] = p
            elif old is not None and p is not None and p < old:
                best_price_by_text[t] = p

    candidates_texts = list(best_price_by_text.keys())
    candidates_prices = [best_price_by_text[t] for t in candidates_texts]

    print(f"Total candidate rows (deduplicated by text): {len(candidates_texts)}")
    return candidates_texts, candidates_prices


# ---------------- MODEL ENCODING ----------------

def encode_texts(texts: List[str],
                 tokenizer: AutoTokenizer,
                 encoder: AutoModel,
                 device: str,
                 max_len: int = MAX_LEN,
                 batch_size: int = ENC_BATCH_SIZE) -> torch.Tensor:
    all_embeddings = []
    for i in range(0, len(texts), batch_size):
        batch_texts = texts[i:i + batch_size]
        enc = tokenizer(
            batch_texts,
            padding=True,
            truncation=True,
            max_length=max_len,
            return_tensors="pt"
        )
        input_ids = enc["input_ids"].to(device)
        attention_mask = enc["attention_mask"].to(device)

        with torch.no_grad():
            outputs = encoder(input_ids=input_ids, attention_mask=attention_mask)
            last_hidden = outputs.last_hidden_state
            mask = attention_mask.unsqueeze(-1)
            mean = (last_hidden * mask).sum(dim=1) / mask.sum(dim=1).clamp(min=1e-9)
            normed = F.normalize(mean, p=2, dim=1)

        all_embeddings.append(normed.cpu())

    if not all_embeddings:
        return torch.empty(0, encoder.config.hidden_size, device=device)

    return torch.cat(all_embeddings, dim=0).to(device)


# ---------------- SELECTION WITH PRICE PREFERENCE ----------------

def pick_best_index_with_price(
    raw_cos: np.ndarray,
    final_scores: np.ndarray,
    prices: List[Optional[float]],
    good_threshold: float = GOOD_SIM_THRESHOLD,
) -> int:
    """
    If there are >=2 "good" candidates (raw cosine >= good_threshold),
    pick the one with the LOWEST price (missing price = expensive),
    tie-break by higher raw cosine.
    Otherwise pick by highest final_scores.
    """
    good_idx = np.where(raw_cos >= good_threshold)[0]
    if len(good_idx) >= 2:
        # sort good candidates by (price asc, raw_cos desc)
        def price_key(j: int):
            p = prices[j]
            p = MISSING_PRICE if p is None else float(p)
            return (p, -float(raw_cos[j]))
        best = min(good_idx, key=price_key)
        return int(best)

    # fallback: best final score
    return int(np.argmax(final_scores))


def pick_second_index(
    best_idx: int,
    raw_cos: np.ndarray,
    final_scores: np.ndarray,
    prices: List[Optional[float]],
) -> int:
    """
    Pick second best with the same logic, excluding best_idx.
    """
    if len(final_scores) <= 1:
        return -1

    mask = np.ones(len(final_scores), dtype=bool)
    mask[best_idx] = False

    raw2 = raw_cos[mask]
    fin2 = final_scores[mask]
    idx_map = np.where(mask)[0]

    second = pick_best_index_with_price(raw2, fin2, [prices[i] for i in idx_map], GOOD_SIM_THRESHOLD)
    return int(idx_map[second])


# ---------------- MAIN ----------------

def main():
    print(f"DEVICE: {DEVICE}")

    print(f"Loading biencoder model from: {MODEL_DIR}")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_DIR)
    encoder = AutoModel.from_pretrained(MODEL_DIR).to(DEVICE)
    encoder.eval()

    # Candidates (text + price)
    candidates, candidate_prices = load_all_candidates_from_dapt(DAPT_DIR)
    if not candidates:
        raise RuntimeError("No candidates found in dapt directory (.xlsx/.xls).")

    print("Encoding candidate texts...")
    candidate_embs = encode_texts(candidates, tokenizer, encoder, DEVICE)
    candidate_params = [extract_params(t) for t in candidates]

    # Samples
    print(f"Loading samples from: {SAMPLES_FILE}")
    df_samples = pd.read_excel(SAMPLES_FILE)
    col_a = df_samples.iloc[:, 0].astype(str).tolist()
    col_b_ref = df_samples.iloc[:, 1].astype(str).tolist() if df_samples.shape[1] >= 2 else [""] * len(col_a)

    print("Encoding sample A texts...")
    sample_embs = encode_texts(col_a, tokenizer, encoder, DEVICE)

    print("Encoding reference B (for green highlight)...")
    ref_embs = encode_texts(col_b_ref, tokenizer, encoder, DEVICE)

    out_rows = []
    ae_sims = []

    for i, (text_a, ref_b) in enumerate(zip(col_a, col_b_ref)):
        q_emb = sample_embs[i]
        q_params = extract_params(text_a)

        with torch.no_grad():
            raw_cos = torch.matmul(q_emb.unsqueeze(0), candidate_embs.t()).cpu().numpy()[0]  # (N,)

        penalties = np.array([parameter_mismatch_penalty(q_params, candidate_params[j]) for j in range(len(candidates))], dtype=np.float32)
        final_scores = raw_cos - penalties

        best_idx = pick_best_index_with_price(raw_cos, final_scores, candidate_prices, GOOD_SIM_THRESHOLD)
        second_idx = pick_second_index(best_idx, raw_cos, final_scores, candidate_prices)

        best_match = candidates[best_idx]
        best_price = candidate_prices[best_idx]
        best_cos = float(raw_cos[best_idx])
        best_pen = float(penalties[best_idx])
        best_final = float(final_scores[best_idx])

        if second_idx >= 0:
            second_match = candidates[second_idx]
            second_price = candidate_prices[second_idx]
            second_final = float(final_scores[second_idx])
            score_diff = best_final - second_final
        else:
            second_match = ""
            second_price = None
            score_diff = float("nan")

        # A–E similarity for green highlighting
        if str(ref_b).strip() == "":
            ae_sim = float("nan")
        else:
            with torch.no_grad():
                ae_sim = float(torch.dot(sample_embs[i], ref_embs[i]).item())
        ae_sims.append(ae_sim)

        out_rows.append({
            "A_sample": text_a,
            "B_best_match": best_match,
            "C_second_match": second_match,
            "D_best_match_similarity": best_cos,      # raw cosine
            "E_param_mismatch": best_pen,             # penalty
            "F_best_price": best_price,               # NEW
            "G_second_price": second_price,           # NEW
            "H_score_diff": score_diff,               # final score diff
            "I_sample_B_ref": ref_b,
            "J_AE_similarity": ae_sim,
        })

        if (i + 1) % 50 == 0:
            print(f"Processed {i+1}/{len(col_a)} rows...")

    output_file = "samples_with_matches.xlsx"
    out_df = pd.DataFrame(out_rows)
    out_df.to_excel(output_file, index=False)
    print(f"Saved: {output_file}")

    # Green highlighting
    wb = load_workbook(output_file)
    ws = wb.active
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for row_idx, sim in enumerate(ae_sims, start=2):
        if sim is None or math.isnan(sim):
            continue
        if sim >= AE_SIM_THRESHOLD:
            for cell in ws[row_idx]:
                cell.fill = green_fill

    wb.save(output_file)
    print(f"Applied green highlighting (A–E sim >= {AE_SIM_THRESHOLD}).")


if __name__ == "__main__":
    main()
