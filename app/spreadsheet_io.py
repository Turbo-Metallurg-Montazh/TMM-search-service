from io import BytesIO
from typing import Any

from fastapi import UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import column_index_from_string, get_column_letter


def _to_rgb_hex(color) -> str | None:
    if not color:
        return None

    rgb = None
    if hasattr(color, "rgb") and isinstance(color.rgb, str):
        rgb = color.rgb
    elif isinstance(color, str):
        rgb = color
    elif hasattr(color, "indexed") and color.indexed is not None:
        try:
            idx = int(color.indexed)
        except Exception:
            return None
        if 0 <= idx < len(COLOR_INDEX):
            rgb = COLOR_INDEX[idx]
    if not rgb:
        return None

    if len(rgb) == 8:
        rgb = rgb[2:]
    return f"#{rgb.lower()}"


def _excel_width_to_px(width: float) -> int:
    return int(round(width * 7 + 5))


async def import_xlsx_to_univer(file: UploadFile) -> dict[str, Any]:
    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=False)

    sheets = []
    for ws in wb.worksheets:
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None and (cell.fill is None or cell.fill.patternType is None):
                    continue

                bg = None
                if cell.fill and cell.fill.patternType:
                    bg = _to_rgb_hex(cell.fill.fgColor)

                cell_obj: dict[str, Any] = {"r": cell.row - 1, "c": cell.column - 1, "v": cell.value}

                style = {}
                if bg:
                    style["bg"] = {"rgb": bg}
                if cell.font and cell.font.bold:
                    style["bl"] = 1
                if style:
                    cell_obj["s"] = style

                if cell.number_format:
                    cell_obj["numberFormat"] = cell.number_format

                cells.append(cell_obj)

        col_widths = {}
        for col_letter, dim in ws.column_dimensions.items():
            if dim.width:
                idx0 = column_index_from_string(col_letter) - 1
                col_widths[str(idx0)] = _excel_width_to_px(dim.width)

        sheets.append({"name": ws.title, "cells": cells, "colWidths": col_widths})

    return {"sheets": sheets}


def export_univer_to_xlsx(payload: dict[str, Any]) -> StreamingResponse:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in payload.get("sheets", []):
        ws = wb.create_sheet(title=str(sheet["name"])[:31])

        for cell in sheet.get("cells", []):
            r = int(cell["r"]) + 1
            c = int(cell["c"]) + 1
            excel_cell = ws.cell(row=r, column=c, value=cell.get("v"))

            if cell.get("bold"):
                excel_cell.font = Font(bold=True)
            if cell.get("numberFormat"):
                excel_cell.number_format = cell["numberFormat"]

        for col, width in sheet.get("colWidths", {}).items():
            ws.column_dimensions[get_column_letter(int(col) + 1)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="export.xlsx"'},
    )

