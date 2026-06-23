from __future__ import annotations

import asyncio
from io import BytesIO

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Color

from app.spreadsheet_io import _excel_width_to_px, _to_rgb_hex, export_univer_to_xlsx, import_xlsx_to_univer


def _xlsx_upload(workbook: Workbook, filename: str = "table.xlsx") -> UploadFile:
    buf = BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return UploadFile(filename=filename, file=buf)


async def _response_body(response) -> bytes:
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk)
    return b"".join(chunks)


def test_to_rgb_hex_normalizes_argb_and_indexed_colors():
    assert _to_rgb_hex("FF112233") == "#112233"
    assert _to_rgb_hex(Color(indexed=2)) is not None
    assert _to_rgb_hex(None) is None


def test_excel_width_to_px_converts_width():
    assert _excel_width_to_px(10) == 75


def test_import_xlsx_to_univer_reads_cells_styles_and_widths():
    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица"
    ws["A1"] = "Название"
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    ws["B2"] = 123.45
    ws["B2"].number_format = "0.00"
    ws.column_dimensions["A"].width = 10

    result = asyncio.run(import_xlsx_to_univer(_xlsx_upload(wb)))

    assert result["sheets"][0]["name"] == "Таблица"
    assert result["sheets"][0]["colWidths"] == {"0": 75}
    assert {
        "r": 0,
        "c": 0,
        "v": "Название",
        "s": {"bg": {"rgb": "#ff0000"}, "bl": 1},
        "numberFormat": "General",
    } in result["sheets"][0]["cells"]
    assert {"r": 1, "c": 1, "v": 123.45, "numberFormat": "0.00"} in result["sheets"][0]["cells"]


def test_export_univer_to_xlsx_returns_workbook_response():
    response = export_univer_to_xlsx(
        {
            "sheets": [
                {
                    "name": "Very long sheet name that must be trimmed",
                    "cells": [
                        {"r": 0, "c": 0, "v": "Название", "bold": True},
                        {"r": 1, "c": 1, "v": 123.45, "numberFormat": "0.00"},
                    ],
                    "colWidths": {"0": 12},
                }
            ]
        }
    )

    body = asyncio.run(_response_body(response))
    workbook = load_workbook(BytesIO(body))
    ws = workbook.active

    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert ws.title == "Very long sheet name that must "
    assert ws["A1"].value == "Название"
    assert ws["A1"].font.bold is True
    assert ws["B2"].value == 123.45
    assert ws["B2"].number_format == "0.00"
    assert ws.column_dimensions["A"].width == 12
